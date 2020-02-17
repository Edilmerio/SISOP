import io
import os
from datetime import datetime

from django.shortcuts import render, redirect, reverse
from django.views.generic import TemplateView
from django.db.models import Subquery, OuterRef, Q
from django.contrib.auth.mixins import LoginRequiredMixin
from django_tables2.export.export import TableExport
from openpyxl import Workbook, load_workbook

from espectro.models import Sistema, Solicitud
from espectro import utiles, tables
from general.models import DivisionTerritorial
from general import utils
from config.settings_base_dir import BASE_DIR
from general.user_metadata import UserPreference


class ListadoSolicitudesLicencias(TemplateView):
    template_name = 'espectro/TablaSolicitudLicencia.html'

    def get(self, request, *args, **kwargs):
        try:
            sistema = Sistema.objects.get(id=args[0])
        except Sistema.DoesNotExist:
            # enviar notificacion a esta pagina.
            return redirect(reverse('espectro:listado sistemas', args={1}), permanent=True)
        tb_sol_page = int(request.GET.get('tb_sol_page', 1))
        table_solicitudes, solicitud_marcada = utiles.tabla_solicitudes_factory(sistema, tb_sol_page)
        table_licencias = utiles.tabla_licencias_factory(solicitud_marcada)
        return render(request, self.template_name, {'table_solicitudes': table_solicitudes,
                                                    'table_licencias': table_licencias, 'sistema': sistema,})


class ListadoLicencias(TemplateView):
    template_name = 'espectro/TablaLicencias.html'

    def get(self, request, *args, **kwargs):
        try:
            solicitud = Solicitud.objects.get(id=args[0])
        except Solicitud.DoesNotExist:
            # enviar notificacion a esta pagina.
            return redirect(reverse('espectro:listado sistemas', args={1}), permanent=True)
        tb_lic_page = int(request.GET.get('tb_lic_page', 1))
        table_licencias = utiles.tabla_licencias_factory(solicitud, tb_lic_page)
        return render(request, self.template_name, {'table_licencias': table_licencias, 'sistema': solicitud.sistema})


class ListadoSistemas(LoginRequiredMixin, TemplateView):
    template_name = 'espectro/ListadoSistemas.html'

    def __init__(self):
        self.per_page_list = ['8', '10', '15', '25', 'todos']
        super().__init__()

    def get(self, request, *args, **kwargs):
        # si args[0]=='1' seran los sistemas que estan en Uso (tinen el campo esta_en_uso == True)
        # si args[0]=='0' seran los sistemas pendientes de instalar (tinen el campo esta_en_uso == False
        # y su ultima solicitud no es de baja)
        notify = request.GET.get('notify', None)

        try:
            div_user = DivisionTerritorial.objects.get(identificativo=request.user.unidad_org.strip())
        except DivisionTerritorial.DoesNotExist:
            div_user = None
        if request.user.has_perm('espectro.visualizador_nacional'):
            sistemas_visualizar = Sistema.objects.all()
        else:
            sistemas_visualizar = Sistema.objects.filter(division_territorial=div_user)
        sistemas_sin_sol_baja = sistemas_visualizar.exclude(solicitud__tipo_solicitud__tipo_solicitud='BAJA')
        if args[0] == '1':
            sistemas_buscados = sistemas_sin_sol_baja.filter(esta_en_uso=True)
        else:
            sistemas_buscados = sistemas_sin_sol_baja.filter(esta_en_uso=False)
        search = request.GET.get('search', '')
        if search != '':
            sistemas_filter_search = sistemas_buscados.filter(
                Q(enlace__icontains=request.GET.get('search')) | Q(sistema__icontains=request.GET.get('search')))
        else:
            sistemas_filter_search = sistemas_buscados
        if request.user.has_perm('espectro.visualizador_nacional'):
            dict_loc = utiles.div_terrotiriales_validas_seleccionadas(request.GET.get('list_mun', 'todos'))
            sistemas_filter_loc = sistemas_filter_search \
                .filter(division_territorial__nombre__in=[s for s in dict_loc if dict_loc[s] == 'selected'])
        else:
            dict_loc = utiles.municipios_validos_seleccionados(div_user, request.GET.get('list_mun', 'todos'))
            sistemas_filter_loc = sistemas_filter_search \
                .filter(radio__municipio__nombre__in=[s for s in dict_loc if dict_loc[s] == 'selected']).distinct()

        table = tables.SistemasTable(sistemas_filter_loc)
        table_order_by = request.GET.get('sort', 'estado_licencia')
        columns_ordered = [v for k, v in table.columns.items()
                           if (v.attrs['th'].get('class') and 'orderable' in v.attrs['th'].get('class'))]
        name_orderable_coloum = [colum.order_by_alias for colum in columns_ordered] + \
                                ['-' + colum.order_by_alias for colum in columns_ordered]
        table_order_by = 'estado_licencia' if table_order_by not in name_orderable_coloum else table_order_by
        table.order_by = table_order_by
        export_format = request.GET.get('_export', None)
        if not export_format:
            cant_elem = sistemas_filter_loc.count()
            per_page = request.GET.get('per_pag', None)
            user_pref = UserPreference(request.user)
            per_page = UserPreference.determine_value_preference(per_page, user_pref.per_page_list_sist_inst, self.per_page_list)[0]
            user_pref.per_page_list_sist_inst = per_page
            int_per_page = cant_elem if per_page == 'todos' else int(per_page)
            page = utiles.pagina_valida(request.GET.get('page', 1), int_per_page, cant_elem)
            table.paginate(page=page, per_page=int_per_page)
            table.cant_rows = len(list(table.data))
            dict_per_page = utiles.dict_elemen_seleced(self.per_page_list, [per_page])
            return render(request, self.template_name,
                          {'table': table, 'busqueda': request.GET.get('search', ''),
                           'municipio_dt_dict': dict_loc, 'dict_per_page': dict_per_page,
                           'cant_elem': cant_elem,
                           'notify': notify})
        else:
            if not TableExport.is_valid_format(export_format):
                export_format = 'xlsx'
            exporter = TableExport(export_format, table, exclude_columns=('editar',))
            wb = load_workbook(filename=io.BytesIO(exporter.export()))
            wb.active.title = 'Sistemas ' + ('en Uso' if args[0] == '1' else 'pte. instalar')
            name_base = 'listado de sistemas {}'.format(datetime.now().strftime("%d-%m-%Y %H-%M-%S"))
            ext_file = export_format
            dir_directory = os.path.join('temp', 'daily')
            utils.make_directory(os.path.join(BASE_DIR, dir_directory))
            name_file = utils.name_file_unique(dir_directory, name_base, ext_file)
            abs_path = os.path.join(BASE_DIR, dir_directory, name_file)
            wb.save(os.path.join(abs_path))
            return redirect(reverse('espectro:descargar_ficheros')+'?dir={}'.format(abs_path), permanent=False)


class ListadoSolicitudesProceso(LoginRequiredMixin, TemplateView):

    """
    Se encarga de mostrar todas las solicitudes en proceso y sistemas sin tramitar, solicitudes en proceso se entiende por toda solicitud
    que no tenga fecha de autorizacion y que no se halla autorizado ninguna despues de esta (para el mismo sistema).
    :arg[0] == 0 todas las solicitudes pendientes.
    :arg[0] == 1 solicitudes en proceso de alta.
    :arg[0] == 2 solicitudes en proceso de baja.
    :arg[0] == 3 solicitudes en proceso de modificacion.
    :arg[0] =! 0,1,2,3 sistemas sin tramitar.
    """
    template_name = 'espectro/ListadoSolicitudesProceso.html'

    def __init__(self):
        self.per_page_list = ['8', '10', '15', '25', 'todos']
        super().__init__()

    def get(self, request, *args, **kwargs):

        per_page = request.GET.get('per_pag', None)
        user_pref = UserPreference(request.user)
        per_page = UserPreference.determine_value_preference(per_page, user_pref.per_page_list_sol_proc, self.per_page_list)[0]
        user_pref.per_page_list_sol_proc = per_page

        dict_per_page = utiles.dict_elemen_seleced(self.per_page_list, [per_page])

        try:
            div_user = DivisionTerritorial.objects.get(identificativo=request.user.unidad_org.strip())
        except DivisionTerritorial.DoesNotExist:
            div_user = None
        search = request.GET.get('search', '')
        if int(args[0]) in [0, 1, 2, 3]:
            solicitudes_autorizadas = Solicitud.objects.exclude(fecha_autorizacion=None)
            ultima_solicitud_autorizada = Subquery(solicitudes_autorizadas
                                                   .filter(sistema__id=OuterRef('sistema__id'), id__gt=OuterRef('id'))
                                                   .order_by('-id').values('id')[:1])
            solicitudes_sin_fecha_autorizacion = Solicitud.objects\
                .filter(fecha_autorizacion=None).annotate(id_ult_sol_aut=ultima_solicitud_autorizada)
            solicitudes_sin_fecha_autorizacion_pte = solicitudes_sin_fecha_autorizacion.filter(id_ult_sol_aut=None)\
                .order_by('sistema_id', '-fecha_envio')
            if args[0] == '0':
                solicitudes_pte = solicitudes_sin_fecha_autorizacion_pte
            elif args[0] == '1':
                solicitudes_pte = solicitudes_sin_fecha_autorizacion_pte.filter(tipo_solicitud__tipo_solicitud='ALTA')
            elif args[0] == '2':
                solicitudes_pte = solicitudes_sin_fecha_autorizacion_pte.filter(tipo_solicitud__tipo_solicitud='BAJA')
            else:
                solicitudes_pte = solicitudes_sin_fecha_autorizacion_pte.filter(tipo_solicitud__tipo_solicitud='MODIFICACION')
            if search != '':
                # busca en las solicitudes pendientes las que el sistema tenga el enlace o el numero
                try:
                    sol_filter_search = solicitudes_pte.filter(sistema__sistema__icontains=int(request.GET.get('search')))
                except ValueError:
                    sol_filter_search = solicitudes_pte.filter(sistema__enlace__icontains=request.GET.get('search'))
            else:
                sol_filter_search = solicitudes_pte

            if request.user.has_perm('espectro.visualizador_nacional'):
                dict_loc = utiles.div_terrotiriales_validas_seleccionadas(request.GET.get('list_mun', 'todos'))
                sol_filter_loc = sol_filter_search\
                    .filter(sistema__division_territorial__nombre__in=[s for s in dict_loc if dict_loc[s] == 'selected'])\
                    .order_by('-fecha_envio').distinct()
            else:
                dict_loc = utiles.municipios_validos_seleccionados(div_user, request.GET.get('list_mun', 'todos'))
                sol_filter_loc = sol_filter_search\
                    .filter(sistema__radio__municipio__nombre__in=[s for s in dict_loc if dict_loc[s] == 'selected'])\
                    .order_by('-fecha_envio').distinct()

            table_sol_pte = tables.SolicitudesProcesoTable(sol_filter_loc)
            export_format = request.GET.get('_export', None)
            if not export_format:
                cant_elem = sol_filter_loc.count()
                int_per_page = cant_elem if per_page == 'todos' else int(per_page)
                page = utiles.pagina_valida(request.GET.get('page', 1), int_per_page, cant_elem)
                table_sol_pte.paginate(page=page, per_page=int_per_page)
                table_sol_pte.cant_rows = len(list(table_sol_pte.data))
                return render(request, self.template_name, {'table': table_sol_pte, 'busqueda': request.GET.get('search', ''),
                                                            'municipio_dt_dict': dict_loc, 'dict_per_page': dict_per_page,
                                                            'cant_elem': cant_elem})
            else:
                if not TableExport.is_valid_format(export_format):
                    export_format = 'xlsx'
                exporter = TableExport(export_format, table_sol_pte, exclude_columns=('editar', 'id', 'archivo_solicitud'))
                wb = load_workbook(filename=io.BytesIO(exporter.export()))
                name_sheet = ['todas', 'alta', 'baja', 'modificacion']
                wb.active.title = 'Solicitudes pte. ' + name_sheet[int(args[0])]
                name_base = 'listado de solicitudes pendientes {}'.format(datetime.now().strftime("%d-%m-%Y %H-%M-%S"))
                ext_file = export_format
                dir_directory = os.path.join('temp', 'daily')
                utils.make_directory(os.path.join(BASE_DIR, dir_directory))
                name_file = utils.name_file_unique(dir_directory, name_base, ext_file)
                abs_path = os.path.join(BASE_DIR, dir_directory, name_file)
                wb.save(os.path.join(abs_path))
                return redirect(reverse('espectro:descargar_ficheros') + '?dir={}'.format(abs_path), permanent=False)

        # Sistemas sin tramitar------------------------------------------------------------------------------------------------------------
        sistemas_pte_tramitar = Sistema.objects.filter(solicitud=None)
        if search != '':
            # busca los sistemas pendientes de tramitar
            try:
                sist_pte_tramitar_search = sistemas_pte_tramitar.filter(sistema__icontains=int(request.GET.get('search')))
            except ValueError:
                sist_pte_tramitar_search = sistemas_pte_tramitar.filter(solicitud=None, enlace__icontains=request.GET.get('search'))
        else:
            sist_pte_tramitar_search = sistemas_pte_tramitar
        if request.user.has_perm('espectro.visualizador_nacional'):
            dict_loc = utiles.div_terrotiriales_validas_seleccionadas(request.GET.get('list_mun', 'todos'))
            sist_filter_loc = sist_pte_tramitar_search\
                .filter(division_territorial__nombre__in=[s for s in dict_loc if dict_loc[s] == 'selected']).distinct()
        else:
            dict_loc = utiles.municipios_validos_seleccionados(div_user, request.GET.get('list_mun', 'todos'))
            sist_filter_loc = sist_pte_tramitar_search \
                .filter(radio__municipio__nombre__in=[s for s in dict_loc if dict_loc[s] == 'selected']).distinct()

        table_sistema_pte_tramitar = tables.SistemasTable(sist_filter_loc)
        table_sistema_pte_tramitar.columns.hide('estado_licencia')
        table_order_by = request.GET.get('sort', '-id')
        columns_ordered = [v for k, v in table_sistema_pte_tramitar.columns.items()
                           if (v.attrs['th'].get('class') and 'orderable' in v.attrs['th'].get('class'))]
        name_orderable_coloum = [colum.order_by_alias for colum in columns_ordered] + \
                                ['-' + colum.order_by_alias for colum in columns_ordered]
        table_order_by = '-id' if table_order_by not in name_orderable_coloum else table_order_by
        table_sistema_pte_tramitar.order_by = table_order_by
        export_format = request.GET.get('_export', None)
        if not export_format:
            cant_elem = sist_filter_loc.count()
            int_per_page = cant_elem if per_page == 'todos' else int(per_page)
            page = utiles.pagina_valida(request.GET.get('page', 1), int_per_page, cant_elem)
            table_sistema_pte_tramitar.paginate(page=page, per_page=int_per_page)
            table_sistema_pte_tramitar.cant_rows = len(list(table_sistema_pte_tramitar.data))
            return render(request, self.template_name, {'table': table_sistema_pte_tramitar, 'busqueda': request.GET.get('search', ''),
                                                        'municipio_dt_dict': dict_loc, 'dict_per_page': dict_per_page,
                                                        'cant_elem': cant_elem})
        else:
            if not TableExport.is_valid_format(export_format):
                export_format = 'xlsx'
            exporter = TableExport(export_format, table_sistema_pte_tramitar, exclude_columns=('editar',))
            wb = load_workbook(filename=io.BytesIO(exporter.export()))
            wb.active.title = 'Solicitudes pte. tramitar'
            name_base = 'sistemas sin tramitar {}'.format(datetime.now().strftime("%d-%m-%Y %H-%M-%S"))
            ext_file = export_format
            dir_directory = os.path.join('temp', 'daily')
            utils.make_directory(os.path.join(BASE_DIR, dir_directory))
            name_file = utils.name_file_unique(dir_directory, name_base, ext_file)
            abs_path = os.path.join(BASE_DIR, dir_directory, name_file)
            wb.save(os.path.join(abs_path))
        return redirect(reverse('espectro:descargar_ficheros') + '?dir={}'.format(abs_path), permanent=False)
