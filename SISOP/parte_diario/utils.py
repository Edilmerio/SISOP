from datetime import datetime, date, timedelta
import string
import os

from django.db.models import Min, Max
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django_tables2.export.export import TableExport
from django.forms import modelformset_factory

from .calcular_indicadores import CalcularIndicadores
from .tables import IndicadoresTable
from .models import Pendientes, LineasServicios
from general import utils
from config.settings_base_dir import BASE_DIR
from .forms import LineasServicioForms
from general.models import CentroAsociado

def diff_days(date1, date2):
    """
    return la diferencia de dias entre las
    2 fechas
    """
    d1 = datetime.strptime(date1, '%Y-%m-%d')
    d2 = datetime.strptime(date2, '%Y-%m-%d')
    return (d2 - d1).days

def fecha_ini_fin_disponible_calculo(dt):
    """
    devulve la fecha inicial y final disponibe para
    el calculo de los indicadores
    fecha_inicial es la fecha mas antigua en BD
    fecha_fin es la fecha mas reciente en BD
    se obtiene de la tabla pendientes.
    """
    pendientes_dt = Pendientes.objects.filter(central__centro_asociado__centro_principal__division_territorial=dt)
    if not pendientes_dt.exists():
        d = date(year=2100, month=1, day=1)
        return d, d
    fecha_inicio = pendientes_dt.aggregate(fecha_inicial=Min('fecha'))['fecha_inicial'] + timedelta(days=1)
    fecha_fin = pendientes_dt.aggregate(fecha_fin=Max('fecha'))['fecha_fin']
    return fecha_inicio, fecha_fin

class DoTableIndicadores:
    def __init__(self, fecha_inicio, fecha_fin,  dt):
        self.dt = dt
        self.fecha_inicio = fecha_inicio
        self.fecha_fin = fecha_fin
        self.indicadores_cta = None
        self.indicadores_ctp = None
        self.indicadores_dt = None
        self.tipo = None
        self.fecha_utilizar_lineas = None
        self.do_calcule()

    def do_calcule(self):
        a = CalcularIndicadores(self.fecha_inicio, self.fecha_fin, self.dt)
        a.obtener_data_primarios()
        self.fecha_utilizar_lineas = a.get_fecha_utilizar_lineas()
        self.indicadores_cta = a.calculo_indicadores_nivel_cta()
        self.indicadores_ctp = a.calculo_indicadores_nivel_ctp()
        self.indicadores_dt = a.calculo_indicadores_nivel_dt()

    def table_nivel_cta(self):
        table_indicadores_cta_tb = IndicadoresTable(self.indicadores_cta[0]+self.indicadores_dt[0])
        table_indicadores_cta_tx = IndicadoresTable(self.indicadores_cta[1]+self.indicadores_dt[1])
        return table_indicadores_cta_tb, table_indicadores_cta_tx

    def table_nivel_ctp(self):
        table_indicadores_ctp_tb = IndicadoresTable(self.indicadores_ctp[0]+self.indicadores_dt[0])
        table_indicadores_ctp_tx = IndicadoresTable(self.indicadores_ctp[1]+self.indicadores_dt[1])
        return table_indicadores_ctp_tb, table_indicadores_ctp_tx

    def do_title_tables(self):
        if self.fecha_fin == self.fecha_inicio:
            self.tipo = 'Diario'
            rango = self.fecha_inicio.strftime('%d-%m-%Y')
        else:
            self.tipo = 'Acumulado'
            rango ='de {} hasta {}'.format(self.fecha_inicio.strftime('%d-%m-%Y'), self.fecha_fin.strftime('%d-%m-%Y'))

        string_fecha_lineas = 'No hay líneas' if not self.fecha_utilizar_lineas else self.fecha_utilizar_lineas.strftime('%m/%Y')

        return 'Indicadores CTA {} Fijo + TFA {} (líneas: {})'.format(self.tipo, rango, string_fecha_lineas),\
               'Indicadores CTP {} Fijo + TFA  {} (líneas: {})'.format(self.tipo, rango, string_fecha_lineas),\
               'Indicadores CTA {} Datos {} (líneas: {})'.format(self.tipo, rango, string_fecha_lineas),\
               'Indicadores CTP {} Datos {} (líneas: {})'.format(self.tipo, rango, string_fecha_lineas)

    def do_titles_short(self):
        return 'CTA Fijo + TFA', 'CTP Fijo + TFA', 'CTA Datos', 'CTP Datos'

    def export_tables_excel(self, table_indicadores_cta_tb, table_indicadores_ctp_tb, table_indicadores_cta_tx, table_indicadores_ctp_tx):
        tables = [table_indicadores_cta_tb, table_indicadores_ctp_tb, table_indicadores_cta_tx, table_indicadores_ctp_tx]
        titles = self.do_title_tables()
        short_title = self.do_titles_short()
        wb = Workbook()
        wb.remove_sheet(wb.active)
        export_format = 'xlsx'
        for t in zip(tables, titles, short_title):
            exporter = TableExport(export_format, t[0])
            ws = wb.create_sheet(t[2])
            ws.append([t[1]])
            ws.merge_cells('A1:O1')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.append(exporter.dataset.headers)
            for d in exporter.dataset.dict:
                ws.append(list(d.values()))
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A1'].font = Font(bold=True, color='FFFFFFFF', size=14)
            ws['A1'].fill = PatternFill(start_color='000066', end_color='000066', fill_type='solid')
            ws.column_dimensions['A'].auto_size = True
            colum_data = string.ascii_uppercase[:ws.max_column]
            for c in colum_data:
                ws[c+'2'].font = Font(bold=True, color='FFFFFFFF')
                ws[c+'2'].fill = PatternFill(start_color='000066', end_color='000066', fill_type='solid')
                ws[c+str(ws.max_row)].font = Font(bold=True, color='FFFFFFFF')
                ws[c+str(ws.max_row)].fill = PatternFill(start_color='000066', end_color='000066', fill_type='solid')
            column_widths = []
            for k, row in enumerate(ws.rows):
                if k == 0:
                    continue
                for i, cell in enumerate(row):
                    if len(column_widths) <= i:
                        column_widths.append(len(str(cell.value)))
                    else:
                        if len(str(cell.value)) > column_widths[i]:
                            column_widths[i] = len(str(cell.value))
            for i, column_width in enumerate(column_widths):
                ws.column_dimensions[get_column_letter(i+1)].width = column_width + 2
        name_base = 'Indicadores {} {}'.format(self.tipo, datetime.now().strftime("%d-%m-%Y %H-%M-%S"))
        ext_file = export_format
        dir_directory = os.path.join('temp', 'daily')
        utils.make_directory(os.path.join(BASE_DIR, dir_directory))
        name_file = utils.name_file_unique(dir_directory, name_base, ext_file)
        abs_path = os.path.join(BASE_DIR, dir_directory, name_file)
        wb.save(abs_path)
        return abs_path


def do_formset_lineas_servicio(dt, fecha):
    lineas_month = LineasServicios.objects \
        .filter(centro_asociado__centro_principal__division_territorial=dt, fecha=fecha) \
        .order_by('centro_asociado__centro_principal__nombre')
    if lineas_month.exists():
        LineasServicioFormSet = modelformset_factory(LineasServicios, form=LineasServicioForms, extra=0)
        lineas_servicio_formset = LineasServicioFormSet(queryset=lineas_month)
    else:
        cta = CentroAsociado.objects.filter(centro_principal__division_territorial=dt) \
            .order_by('centro_principal__nombre')
        cant_cta = cta.count()
        LineasServicioFormSet = modelformset_factory(LineasServicios, form=LineasServicioForms, extra=cant_cta)
        lineas_servicio_formset = LineasServicioFormSet(queryset=lineas_month)
        for t in zip(lineas_servicio_formset, cta):
            t[0].fields['abonado'].initial = 0
            t[0].fields['datos'].initial = 0
            t[0].fields['fecha'].initial = fecha
            t[0].fields['centro_asociado'].initial = t[1].pk
            t[0].cta = t[1].nombre
    return lineas_servicio_formset
